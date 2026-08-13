"""
ICV Evaluation Agent — ADNOC Produced Water Treatment Package
==============================================================
Reviews supplier bids for In-Country Value (ICV) per RFP Section 6.4.

Formula: ICV Score = 15 × min(certified ICV %, 60) ÷ 60, rounded to 2 dp.
ICV weight: 15 / 100 points. Requires valid MoIAT-certified ICV certificate (D6).
"""

import re
import os
import glob
import subprocess
from typing import Optional

# ── Data directory ──────────────────────────────────────────────────────────
DEFAULT_DATA_DIR = os.path.expanduser(
    "~/Downloads/ADNOC_Procurement_Evaluation_Agent/data"
)

# Clean supplier-name mapping from bid filenames
BID_TO_NAME = {
    "Bid_V01_Al_Manara_Process_Solutions_LLC.pdf": "Al Manara Process Solutions LLC",
    "Bid_V02_Gulfstream_Engineering_and_Contracting_WLL.pdf": "Gulfstream Engineering & Contracting WLL",
    "Bid_V03_Rheinwasser_Technik_GmbH.pdf": "Rheinwasser Technik GmbH",
    "Bid_V04_Hanseong_Water_and_Energy_Co_Ltd.pdf": "Hanseong Water & Energy Co Ltd",
    "Bid_V05_Al_Dhafra_Industrial_Services_Co.pdf": "Al Dhafra Industrial Services Co",
    "Bid_V06_Qasr_Al_Bahr_Marine_and_Process_LLC.pdf": "Qasr Al Bahr Marine & Process LLC",
    "Bid_V07_Petrotech_Arabia_Ltd.pdf": "Petrotech Arabia Ltd",
    "Bid_V08_Emirates_Flowline_Systems_FZE.pdf": "Emirates Flowline Systems FZE",
    "Bid_V09_Nakheel_Oilfield_Supplies_and_Services_LLC.pdf": "Nakheel Oilfield Supplies & Services LLC",
    "Bid_V10_Bin_Sultan_Heavy_Industries_LLC.pdf": "Bin Sultan Heavy Industries LLC",
    "Bid_V11_Aquapure_Gulf_Technologies_FZ-LLC.pdf": "Aquapure Gulf Technologies FZ-LLC",
    "Bid_V12_Levant_Energy_Solutions_SAL.pdf": "Levant Energy Solutions SAL",
}

BID_DUE_DATE = "2026-07-16"  # Bid submission date per tracker


# ═══════════════════════════════════════════════════════════════════════════════
#  Data-loading helpers
# ═══════════════════════════════════════════════════════════════════════════════

def pdf_to_text(path: str) -> str:
    """Extract text from a PDF via pdftotext or pypdf fallback."""
    try:
        res = subprocess.run(
            ["pdftotext", "-layout", path, "-"],
            capture_output=True, text=True, timeout=30,
        )
        if res.returncode == 0:
            return res.stdout
    except FileNotFoundError:
        pass
    try:
        import pypdf
        with open(path, "rb") as f:
            reader = pypdf.PdfReader(f)
            return "\n".join(p.extract_text() for p in reader.pages)
    except ImportError:
        raise RuntimeError("Need pdftotext (poppler-utils) or pypdf")


def parse_tracker(data_dir: str) -> list[dict]:
    """Parse the Bid Submission Tracker xlsx."""
    import openpyxl
    path = os.path.join(data_dir, "Bid_Submission_Tracker_2026-0412.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Bid Register"] if "Bid Register" in wb.sheetnames else wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            continue
        if row[0] and str(row[0]).startswith("BID-"):
            rows.append({
                "bid_ref": str(row[0]),
                "company": str(row[1] or ""),
                "country": str(row[2] or ""),
                "city": str(row[3] or ""),
                "contact_person": str(row[4] or ""),
                "email": str(row[5] or ""),
                "phone": str(row[6] or ""),
                "submission_date": str(row[7] or ""),
                "currency": str(row[8] or ""),
                "volumes": str(row[9] or ""),
                "remarks": str(row[10] or ""),
            })
    return rows


def clean_name(filename: str) -> str:
    """Map bid filename to clean supplier name."""
    for k, v in BID_TO_NAME.items():
        if k in filename or filename.endswith(k):
            return v
    name = os.path.basename(filename).replace(".pdf", "")
    name = re.sub(r"Bid_V\d{2}_", "", name).replace("_", " ")
    return name


def parse_supplier(text: str, filename: str, tracker_rows: list[dict]) -> dict:
    """Parse a single supplier bid PDF into structured data."""
    s = {"file": filename, "doc_id": os.path.basename(filename)}
    text_nf = text.replace("\f", "\n")

    def rfind(pattern, group=1, flags=re.I | re.DOTALL):
        m = re.search(pattern, text_nf, flags)
        return m.group(group) if m else None

    # ── ICV data ──────────────────────────────────────────────────────────
    # Certificate line: "Certified score 52% — certificate no. ICV-0184-2025, valid until 31 Mar 2027"
    icv_m = re.search(
        r"Certified\s*score\s+(\d+)%\s*[—–-]?\s*certificate\s*no\.?\s*([\w-]+)",
        text_nf, re.I,
    )
    if not icv_m:
        icv_m = re.search(r"Certified\s*score\s+(\d+)%", text_nf, re.I)
    s["icv_pct"] = int(icv_m.group(1)) if icv_m else None
    s["icv_cert_no"] = (
        icv_m.group(2).strip()
        if icv_m and len(icv_m.groups()) >= 2 and icv_m.group(2)
        else None
    )

    # Certificate validity / expiry
    if re.search(r"Under\s*renewal", text_nf, re.I):
        s["icv_status"] = "under renewal"
        s["icv_valid_until"] = None
    elif s["icv_pct"] is not None:
        s["icv_status"] = "valid"
        # Extract valid-until date
        vm = re.search(r"valid\s*until\s+([\d]+\s+[A-Za-z]+\s+[\d]{4})", text_nf, re.I)
        s["icv_valid_until"] = vm.group(1).strip() if vm else None
    else:
        s["icv_status"] = "not found"
        s["icv_valid_until"] = None

    # ── Issuing body ──────────────────────────────────────────────────────
    # "ICV certificate (MoIAT)"
    ibm = re.search(r"ICV\s*certificate\s*\(([^)]+)\)", text_nf, re.I)
    s["icv_issuing_body"] = ibm.group(1).strip() if ibm else None

    # ── Checklist D6 status ───────────────────────────────────────────────
    s["checklist_d6"] = "N/A"
    sec = None
    for m in re.finditer(
        r"S\s*U\s*B\s*M\s*I\s*S\s*S\s*I\s*O\s*N\s*C\s*H\s*E\s*C\s*K\s*L\s*I\s*S\s*T",
        text_nf, re.I,
    ):
        tail = text_nf[m.end():m.end() + 500]
        if "D1" in tail or "Code" in tail:
            sec = re.match(r"[\s\S]*", text_nf[m.start():m.start() + 3000])
            break
    if sec:
        lines = sec.group(0).split("\n")
        for i, line in enumerate(lines):
            line_s = line.strip()
            m = re.match(r"\s*D(6)\s+", line_s, re.I)
            if not m:
                continue
            # Status on same line after 5+ spaces
            parts = re.split(r"\s{5,}", line_s)
            if len(parts) >= 2:
                candidate = parts[-1].strip()
                if re.match(
                    r"(Enclosed|Under renewal|To be arranged|Not submitted)",
                    candidate, re.I,
                ):
                    s["checklist_d6"] = candidate
                    break
            # Status on previous line
            if i > 0:
                prev = lines[i - 1]
                m2 = re.search(
                    r"\s{8,}(Under\s*renewal[^\n]{0,80})", prev, re.I,
                )
                if m2:
                    s["checklist_d6"] = m2.group(1).strip()
                    break
            # Status on next lines
            for j in range(1, 4):
                if i + j < len(lines):
                    nxt = lines[i + j].strip()
                    m2 = re.match(
                        r"(Enclosed|Under\s*renewal[^\n]{0,80}|To\s*be\s*arranged[^\n]{0,80}|Not\s*submitted)",
                        nxt, re.I,
                    )
                    if m2:
                        s["checklist_d6"] = m2.group(1).strip()
                        break

    # ── Match tracker for country, bid ref ────────────────────────────────
    company_name = clean_name(filename)
    for tr in tracker_rows:
        cname = company_name.lower().replace("&", "and")[:30]
        tname = tr["company"].lower().replace("&", "and")[:30]
        if cname[:20] in tname or tname[:20] in cname:
            s["country"] = tr["country"]
            s["city"] = tr["city"]
            s["bid_ref"] = tr["bid_ref"]
            s["submission_date"] = tr["submission_date"]
            break
    s.setdefault("country", "")
    s.setdefault("city", "")
    s.setdefault("bid_ref", "")

    # ── Local content / fabrication hints ──────────────────────────────────
    # Careful: "ADNOC-LCIG/RFP" contains "LCIG" — avoid matching the tender ref.
    # Only flag *supplier-specific* claims; generic scope boilerplate
    # ("design, fabrication, installation and commissioning") is not evidence.
    s["local_content_hints"] = ""
    if re.search(r"fabricat(?:ion|e)\b[^\n]{0,80}\bUAE\s+facilit", text_nf, re.I | re.DOTALL):
        s["local_content_hints"] = "States fabrication will be performed at its own UAE facility"
    elif re.search(r"fabrication\s+complex[^\n]{0,60}\b(KEZAD|Taweelah|ICAD|Jebel Ali)\b", text_nf, re.I):
        s["local_content_hints"] = "States fabrication complex in UAE free zone (local manufacturing base)"
    elif re.search(r"\bLocal\s+Content\s+Initiative\b", text_nf, re.I):
        s["local_content_hints"] = "Mentions Local Content Initiative (LCI) participation"

    return s


# ═══════════════════════════════════════════════════════════════════════════════
#  ICV Agent
# ═══════════════════════════════════════════════════════════════════════════════

def calculate_icv_score(icv_pct: Optional[float]) -> float:
    """ICV Score = 15 × min(ICV%, 60) ÷ 60, rounded to 2 dp."""
    if icv_pct is None:
        return 0.0
    return round(15 * min(icv_pct, 60) / 60, 2)


def classify_icv_strength(icv_pct: Optional[float], icv_status: str) -> str:
    """Classify ICV Strength: Strong / Medium / Weak / Unknown."""
    if icv_status == "not found" or icv_pct is None:
        return "Unknown"
    if icv_status == "under renewal":
        return "Unknown"
    if icv_pct >= 45:
        return "Strong"
    if icv_pct >= 25:
        return "Medium"
    if icv_pct > 0:
        return "Weak"
    return "Unknown"


def classify_icv_risk(
    icv_pct: Optional[float],
    icv_status: str,
    checklist_d6: str,
) -> str:
    """Classify ICV Risk: Low / Medium / High."""
    # Missing / invalid certificate = High
    if icv_status == "not found" or icv_status == "under renewal":
        return "High"
    if icv_status != "valid" or icv_pct is None:
        return "High"
    # D6 not enclosed
    if "enclosed" not in checklist_d6.lower():
        return "High"
    # Low ICV = Medium
    if icv_pct < 20:
        return "Medium"
    # Moderate but still low
    if icv_pct < 30:
        return "Medium"
    return "Low"


def identify_missing_info(rec: dict) -> list[str]:
    """List missing ICV-related information."""
    missing = []
    if rec.get("icv_pct") is None:
        missing.append("Certified ICV percentage not provided")
    if rec.get("icv_cert_no") is None:
        missing.append("ICV certificate number not provided")
    if rec.get("icv_valid_until") is None and rec.get("icv_status") != "under renewal":
        missing.append("ICV certificate expiry date not provided")
    if rec.get("icv_issuing_body") is None:
        missing.append("ICV issuing body not specified")
    if "enclosed" not in rec.get("checklist_d6", "").lower():
        missing.append("D6 (ICV certificate) not enclosed in submission")
    return missing


def icv_explanation(rec: dict) -> str:
    """Generate a 1–2 sentence explanation per supplier."""
    pct = rec.get("icv_pct")
    status = rec.get("icv_status", "")
    score = calculate_icv_score(pct)
    cert_no = rec.get("icv_cert_no") or "N/A"
    country = rec.get("country", "")

    if status == "under renewal":
        return (
            f"ICV certificate is under renewal (not valid on bid due date "
            f"{BID_DUE_DATE}). Scores 0/15 and fails mandatory D6 screening. "
            f"Supplier is based in {country}."
        )
    if pct is None:
        return (
            f"No valid ICV certificate found. Scores 0/15 and fails "
            f"mandatory D6 screening. Supplier is based in {country}."
        )
    strength = classify_icv_strength(pct, status)
    risk = classify_icv_risk(pct, status, rec.get("checklist_d6", ""))
    explanation = (
        f"Certified ICV score of {pct}% (certificate {cert_no}, "
        f"MoIAT-certified, valid until {rec.get('icv_valid_until', 'not specified')}). "
        f"Achieves {score}/15 ICV points. "
        f"ICV position is {strength.lower()} with {risk.lower()} risk. "
        f"Supplier is based in {country}."
    )
    if rec.get("local_content_hints"):
        explanation += f" {rec['local_content_hints']}."
    return explanation


def evaluate_supplier(rec: dict) -> dict:
    """Evaluate one supplier per the ICV Agent brief format."""
    pct = rec.get("icv_pct")
    status = rec.get("icv_status", "not found")
    cert_no = rec.get("icv_cert_no") or "N/A"
    checklist_d6 = rec.get("checklist_d6") or "N/A"

    # ICV Certificate status
    if status == "valid":
        cert_status = "Valid"
    elif status == "under renewal":
        cert_status = "Invalid"
    else:
        cert_status = "Missing"

    score = calculate_icv_score(pct)
    strength = classify_icv_strength(pct, status)
    missing = identify_missing_info(rec)
    risk = classify_icv_risk(pct, status, checklist_d6)
    explanation = icv_explanation(rec)

    return {
        "supplier_name": clean_name(rec.get("file", "")),
        "icv_certificate": cert_status,
        "certified_icv_score": f"{pct}%" if pct is not None else "Not provided",
        "icv_points": score,
        "icv_points_str": f"{score:.2f}/15",
        "icv_strength": strength,
        "missing_information": missing if missing else ["None"],
        "icv_risk": risk,
        "explanation": explanation,
        # Raw data for comparison table
        "icv_pct": pct,
        "cert_no": cert_no,
        "country": rec.get("country", ""),
        "checklist_d6": checklist_d6,
        "valid_until": rec.get("icv_valid_until", "N/A"),
        "local_content_hints": rec.get("local_content_hints", ""),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  Report generation
# ═══════════════════════════════════════════════════════════════════════════════

def generate_comparison_table(results: list[dict]) -> str:
    """Generate a formatted ICV comparison table."""
    lines = []
    lines.append("=" * 120)
    lines.append("ICV COMPARISON TABLE — ADNOC Produced Water Treatment Package")
    lines.append(f"Tender: ADNOC-LCIG/RFP/2026-0412  |  ICV Weight: 15/100 points")
    lines.append("=" * 120)
    lines.append("")
    lines.append(
        f"{'Supplier':<38s} {'ICV%':>6s} {'Cert No':<18s} {'Status':<10s} "
        f"{'Points':>7s} {'Strength':<10s} {'Risk':<8s}"
    )
    lines.append("-" * 120)
    for r in results:
        lines.append(
            f"{r['supplier_name']:<38s} {r['certified_icv_score']:>6s} "
            f"{r['cert_no']:<18s} {r['icv_certificate']:<10s} "
            f"{r['icv_points_str']:>7s} {r['icv_strength']:<10s} {r['icv_risk']:<8s}"
        )
    lines.append("-" * 120)
    lines.append("")
    return "\n".join(lines)


def identify_strongest_position(results: list[dict]) -> str:
    """Identify the supplier with the strongest publicly supported ICV position."""
    # Filter to suppliers with valid certificates
    valid = [r for r in results if r["icv_certificate"] == "Valid"]
    if not valid:
        return "No supplier has a valid ICV certificate."

    # Highest ICV% is the strongest
    strongest = max(valid, key=lambda r: r["icv_pct"] or 0)
    # Also check UAE-based suppliers with strong ICV + local content
    uae_valid = [r for r in valid if "United Arab Emirates" in r["country"]]
    best_uae = max(uae_valid, key=lambda r: r["icv_pct"] or 0) if uae_valid else None

    parts = [
        f"**{strongest['supplier_name']}** has the strongest ICV position "
        f"with a certified ICV score of {strongest['certified_icv_score']} "
        f"(certificate {strongest['cert_no']}), achieving {strongest['icv_points_str']} points "
        f"({strongest['icv_strength']} strength, {strongest['icv_risk']} risk)."
    ]
    if best_uae and best_uae["supplier_name"] != strongest["supplier_name"]:
        parts.append(
            f"Among UAE-based suppliers, **{best_uae['supplier_name']}** "
            f"leads with {best_uae['certified_icv_score']} ICV "
            f"({best_uae['icv_points_str']} points)."
        )
    if strongest["local_content_hints"]:
        parts.append(f"Note: {strongest['supplier_name']} {strongest['local_content_hints'].lower()}.")
    return " ".join(parts)


def generate_report(results: list[dict]) -> str:
    """Generate the full ICV evaluation report."""
    lines = []
    lines.append("=" * 72)
    lines.append("ICV EVALUATION REPORT")
    lines.append("Tender: ADNOC-LCIG/RFP/2026-0412 — Produced Water Treatment Package")
    lines.append("=" * 72)
    lines.append("")
    lines.append(f"Total suppliers evaluated: {len(results)}")
    valid_count = sum(1 for r in results if r["icv_certificate"] == "Valid")
    invalid_count = sum(1 for r in results if r["icv_certificate"] == "Invalid")
    missing_count = sum(1 for r in results if r["icv_certificate"] == "Missing")
    lines.append(
        f"  Valid ICV certificates: {valid_count}  |  "
        f"Invalid: {invalid_count}  |  Missing: {missing_count}"
    )
    lines.append("")

    # Per-supplier details
    for r in results:
        lines.append("-" * 72)
        lines.append(f"Supplier Name: {r['supplier_name']}")
        lines.append(f"ICV Certificate: {r['icv_certificate']}")
        lines.append(f"Certified ICV Score: {r['certified_icv_score']}")
        lines.append(f"ICV Points: {r['icv_points_str']}")
        lines.append(f"ICV Strength: {r['icv_strength']}")
        lines.append(f"Missing Information: {'; '.join(r['missing_information'])}")
        lines.append(f"ICV Risk: {r['icv_risk']}")
        lines.append(f"Explanation: {r['explanation']}")
        lines.append("")

    # Comparison table
    lines.append(generate_comparison_table(results))
    lines.append("")

    # Strongest position
    lines.append("STRONGEST PUBLICLY SUPPORTED ICV POSITION")
    lines.append("-" * 72)
    lines.append(identify_strongest_position(results))
    lines.append("")
    lines.append("-" * 72)
    lines.append(
        "Note: This evaluation is based on the provided challenge dataset. "
        "The ICV result is passed to the main procurement agent along with "
        "Technical, Commercial, and HSE results. The final procurement "
        "decision rests with the human procurement engineer."
    )
    lines.append("")
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
#  Main entry point
# ═══════════════════════════════════════════════════════════════════════════════

def run_icv_agent(data_dir: str = DEFAULT_DATA_DIR) -> list[dict]:
    """Run the ICV agent over the challenge dataset. Returns per-supplier results."""
    # Load data
    print(f"[ICV Agent] Loading dataset from {data_dir}...")
    rfp_path = os.path.join(data_dir, "RFP_ADNOC-LCIG_2026-0412_Produced_Water_Treatment.pdf")
    if not os.path.exists(rfp_path):
        raise FileNotFoundError(f"RFP not found at {rfp_path}")

    tracker_rows = parse_tracker(data_dir)
    print(f"  Tracker: {len(tracker_rows)} rows")

    bid_dir = os.path.join(data_dir, "bids")
    suppliers = []
    for fpath in sorted(glob.glob(os.path.join(bid_dir, "*.pdf"))):
        text = pdf_to_text(fpath)
        rec = parse_supplier(text, fpath, tracker_rows)
        suppliers.append(rec)
        print(f"  {clean_name(fpath):<42s} icv={rec.get('icv_pct')} status={rec.get('icv_status')}")

    # Evaluate
    results = [evaluate_supplier(s) for s in suppliers]
    # Sort by ICV points descending
    results.sort(key=lambda r: r["icv_points"], reverse=True)
    return results


def main():
    """CLI entry point."""
    import argparse
    parser = argparse.ArgumentParser(description="ICV Evaluation Agent")
    parser.add_argument(
        "--data-dir", default=DEFAULT_DATA_DIR,
        help=f"Challenge dataset directory (default: {DEFAULT_DATA_DIR})",
    )
    parser.add_argument(
        "--output", "-o", default=None,
        help="Write report to file (default: stdout)",
    )
    args = parser.parse_args()

    results = run_icv_agent(args.data_dir)
    report = generate_report(results)

    if args.output:
        with open(args.output, "w") as f:
            f.write(report)
        print(f"\n[ICV Agent] Report written to {args.output}")
    else:
        print(report)

    # Also return structured results via JSON
    import json
    json_path = os.path.join(
        os.path.dirname(args.output or "."),
        "icv_results.json",
    )
    with open(json_path, "w") as f:
        json.dump(results, f, indent=2)
    print(f"[ICV Agent] Structured results saved to {json_path}")


if __name__ == "__main__":
    main()